import os.path
import pyodbc
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, colors, Font, Alignment
import pyperclip
import sys
import re
import databaseUtility

def printUsage():
    print('Usage:\nCopy the SQL script needs to run, and execute "python databaseExport.py db=db_name_in_config')
    exit()

#Read configuration info
if not os.path.exists('AppSettings.xml'):
    print("ERROR: Cannot find configuration file.")
    exit()

dbConnection = None

#Check argument
databaseNames = ''
if len(sys.argv) < 2:
    printUsage()
else:
    dbNameRe = re.compile(r'^db=(?P<dbName>.+)$')
    dbNameMatch = dbNameRe.match(sys.argv[1])
    if dbNameMatch is not None:
        databaseNames = dbNameMatch.group('dbName').split(',')
    else:
        printUsage()

#Check SQL script
sql = pyperclip.paste()
if len(sql) == 0:
    print('ERROR: Please copy SQL script to clipboard before execution.')
    exit()

print('About to execute:\n{sql}\n'.format(sql = sql))

#Create Excel file
fileName = 'dataExtract.xlsx'
wb = openpyxl.workbook.Workbook()

#Read Style configuration
headerRowDisplayStyle = {'fontColor':'FFFFFF', 'backgroundColor':'000080'}
highLightRowDisplayStyle = {'fontColor':'FFFFFF', 'backgroundColor':'FFFF00'}

configurationContent = '\n'.join(open('AppSettings.xml').readlines())
configurationDocument = BeautifulSoup(configurationContent, features='xml')
uiConfiguration = configurationDocument.find('uiSettings')
if uiConfiguration is not None:
    headerRowConfiguration = uiConfiguration.find('headerRow')
    if headerRowConfiguration is not None:
        if headerRowConfiguration.get('fontColor') is not None:
            headerRowDisplayStyle['fontColor'] = headerRowConfiguration.get('fontColor')
        if headerRowConfiguration.get('backgroundColor') is not None:
            headerRowDisplayStyle['backgroundColor'] = headerRowConfiguration.get('backgroundColor')
    highLightRowConfiguration = uiConfiguration.find('highLightRow')
    if highLightRowConfiguration is not None:
        if highLightRowConfiguration.get('fontColor') is not None:
            highLightRowDisplayStyle['fontColor'] = highLightRowConfiguration.get('fontColor')
        if highLightRowConfiguration.get('backgroundColor') is not None:
            highLightRowDisplayStyle['backgroundColor'] = highLightRowConfiguration.get('backgroundColor')

headerRowFill = PatternFill(start_color=headerRowDisplayStyle['backgroundColor'],
                    end_color=headerRowDisplayStyle['backgroundColor'],
                    fill_type='solid')
highLightRowFill = PatternFill(start_color=highLightRowDisplayStyle['backgroundColor'],
                    end_color=highLightRowDisplayStyle['backgroundColor'],
                    fill_type='solid')
headerRowFont = Font(color=headerRowDisplayStyle['fontColor'])
highLightRowFont = Font(color=highLightRowDisplayStyle['fontColor'])

for databaseName in databaseNames:
    print('Fetching data from {db}\n'.format(db = databaseName))
    
    dbConnection = databaseUtility.getDatabaseConnection(databaseName)
    dbCursor = dbConnection.cursor()
    dbCursor.execute(sql)

    dataSheet = wb.create_sheet(title = databaseName, index = 0)

    #Write Header row
    columnCount = len(dbCursor.description)
    headerIndex = 1
    for header in dbCursor.description:
        headerCell = dataSheet.cell(row = 1, column = headerIndex)
        headerCell.value = header[0]
        headerCell.fill = headerRowFill
        headerCell.font = headerRowFont
        headerIndex += 1

    rowIndex = 2
    for row in dbCursor.fetchall():
        for i in range(0, columnCount):
            dataSheet.cell(row = rowIndex, column = i + 1).value = row[i]
        rowIndex += 1

    cornerCell = dataSheet['A2']
    dataSheet.freeze_panes = cornerCell
    
    dbConnection.close()

#Save SQL script
wrapAlignment = Alignment(wrapText = False, shrinkToFit = False, horizontal = 'left', vertical = 'top')
sqlSheet = wb.create_sheet(title='SQL', index=len(wb.worksheets)-1)
lineIndex = 1
findTabsRegEx = r'^\t+'
for sqlLine in sql.split('\n'):
    tabsCount = 0
    tabPrefixMatch = re.search(findTabsRegEx, sqlLine)
    if tabPrefixMatch is not None:
        tabsCount = len(tabPrefixMatch.group(0))
    sqlSheet.cell(row = lineIndex, column = 1 + tabsCount).alignment = wrapAlignment
    sqlSheet.cell(row = lineIndex, column = 1 + tabsCount).value = sqlLine
    lineIndex += 1

wb.save(fileName)
print('Data saved as ' + fileName)

