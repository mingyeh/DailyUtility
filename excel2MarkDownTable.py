from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pyperclip

class Column:
    captain = ''
    align = 'left'
    maxLength = 0

    def __init__(self, captain, align, maxLength):
        self.captain = captain
        self.align = align
        self.maxLength = maxLength

    def __str__(self):
        return 'Captain:{captain}; Align:{align}; MaxLength:{maxLength}'.format(captain = self.captain, align = self.align, maxLength = self.maxLength)

#Read data from Excel file
wb = load_workbook('sample_data.xlsx')
activeSheet = wb.active

columnCount = activeSheet.max_column
rowCount = activeSheet.max_row

columns = []
for i in range(1, columnCount + 1):
    maxLength = 0
    for j in range(1, rowCount + 1):
        length = len(str(activeSheet.cell(row = j, column = i).value))
        maxLength = length if length > maxLength else maxLength
    columns.append(Column(activeSheet.cell(row = 1, column = i).value,
                          activeSheet.cell(row = 2, column = i).alignment.horizontal,
                          maxLength))

result = '|'
for c in columns:
    result += ' {column} |'.format(column = c.captain.ljust(c.maxLength))

result += '\n'
result += '|'
for c in columns:
    if c.align == 'center':
        result += ':{line}:|'.format(line = '-' * c.maxLength)
    elif c.align == 'right':
        result += ' {line}:|'.format(line = '-' * c.maxLength)
    else:
        result += ' {line} |'.format(line = '-' * c.maxLength)

for r in range(2, rowCount + 1):
    result += '\n'
    result += '|'
    for column in columns:
        columnIndex = columns.index(column) + 1
        cellValue = str(activeSheet.cell(row = r, column = columnIndex).value)
        cellValue = '' if cellValue == 'None' else cellValue
        result += ' {value} |'.format(value = cellValue.rjust(column.maxLength) if column.align == 'right' else cellValue.ljust(column.maxLength))  

wb.close()

print(result)
pyperclip.copy(result)
print('\nResult copied to ClipBoard.')
